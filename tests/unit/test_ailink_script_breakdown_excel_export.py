"""Tests for Script-to-Production Breakdown Excel export."""

from __future__ import annotations

import sys
from pathlib import Path

import pytest

# Ensure src is in path
SRC_DIR = Path(__file__).resolve().parent.parent.parent / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from openpyxl import load_workbook

from ailink_tools.script_breakdown.demo_parser import parse_demo_script
from ailink_tools.script_breakdown.excel_export import (
    ALLOWED_EXTENSIONS,
    export_excel,
)
from ailink_tools.script_breakdown.schemas import BreakdownResult

# ---------------------------------------------------------------------------
# Demo script fixture
# ---------------------------------------------------------------------------

DEMO_SCRIPT = """
ESCENA 1. INT. CASA DE ANA - NOCHE

ANA BRUMA (30 anos) esta sentada en la cocina.
LEO PRADO (25 anos) entra con prisa.

LEO: Han llamado del ayuntamiento.

ESCENA 2. EXT. CAMINO RURAL - AMANECER

Ana y Leo caminan por un camino de tierra.

ANA: El informe dice que necesitamos tres permisos.

ESCENA 3. EXT. PUEBLO DEMO BRUMA - D\xcdA

Ana entra en el BAR DEMO ESTACI\xd3N.

MARA: Ana, ya has hablado con el alcalde?

ESCENA 4. INT. BAR DEMO ESTACI\xd3N - NOCHE

Reunion de produccion.

BRUNO: El sonido va a ser complicado.

ESCENA 5. EXT. CAMPO DE LAVANDA - D\xcdA

Ana camina sola por un campo de lavanda.

ESCENA 6. EXT. CASA DE ANA - D\xcdA

Leo esta reparando una valla.

LEO: Y si el alcalde dice que no?

ESCENA 7. INT. CASA DE ANA - NOCHE

Ana esta sola en la cocina.

ANA: (al telefono) Necesito hablar sobre los permisos.

ESCENA 8. EXT. CAMINO RURAL - NOCHE

Ana y Leo caminan por el camino de noche.

ANA: Manana tenemos que decidir.
"""


@pytest.fixture
def breakdown_result():
    """Parse demo script and return result."""
    return parse_demo_script(DEMO_SCRIPT)


@pytest.fixture
def excel_path(tmp_path):
    """Return Excel output path in tmp_path."""
    return tmp_path / "breakdown.xlsx"


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_export_excel_creates_file(breakdown_result, excel_path):
    """Excel export must create the file."""
    result = export_excel(breakdown_result, excel_path)
    assert result == excel_path
    assert excel_path.exists()


def test_export_excel_has_10_sheets(breakdown_result, excel_path):
    """Excel must have exactly 10 sheets."""
    export_excel(breakdown_result, excel_path)
    wb = load_workbook(str(excel_path))
    assert len(wb.sheetnames) == 10
    wb.close()


def test_export_excel_sheet_names(breakdown_result, excel_path):
    """Excel sheet names must match expected list."""
    export_excel(breakdown_result, excel_path)
    wb = load_workbook(str(excel_path))
    expected = [
        "Resumen", "Escenas", "Personajes", "Localizaciones",
        "Riesgos", "Viabilidad", "Presupuesto", "Recomendaciones",
        "Revisión humana", "Metadata",
    ]
    assert wb.sheetnames == expected
    wb.close()


def test_export_excel_resumen_content(breakdown_result, excel_path):
    """Resumen sheet must have project title and viability."""
    export_excel(breakdown_result, excel_path)
    wb = load_workbook(str(excel_path))
    ws = wb["Resumen"]
    # Find title row
    found_title = False
    found_viability = False
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if row[1] == "Proyecto Demo Bruma":
            found_title = True
        if row[0] == "Viabilidad global":
            found_viability = True
    assert found_title, "Missing project title in Resumen"
    assert found_viability, "Missing viability in Resumen"
    wb.close()


def test_export_excel_escenas_row_count(breakdown_result, excel_path):
    """Escenas sheet must have 8 data rows + header."""
    export_excel(breakdown_result, excel_path)
    wb = load_workbook(str(excel_path))
    ws = wb["Escenas"]
    assert ws.max_row == 9  # 1 header + 8 data rows
    wb.close()


def test_export_excel_personajes_row_count(breakdown_result, excel_path):
    """Personajes sheet must have 5 data rows + header."""
    export_excel(breakdown_result, excel_path)
    wb = load_workbook(str(excel_path))
    ws = wb["Personajes"]
    assert ws.max_row == 6  # 1 header + 5 data rows
    wb.close()


def test_export_excel_presupuesto_total_formula(breakdown_result, excel_path):
    """Presupuesto TOTAL row must have SUM formulas."""
    export_excel(breakdown_result, excel_path)
    wb = load_workbook(str(excel_path))
    ws = wb["Presupuesto"]
    total_row = ws.max_row
    # Check TOTAL label
    assert ws.cell(row=total_row, column=1).value == "TOTAL"
    # Check SUM formulas in columns 3, 4, 5
    for col in [3, 4, 5]:
        cell = ws.cell(row=total_row, column=col)
        assert cell.value is not None, f"Missing formula in col {col}"
        assert str(cell.value).startswith("="), f"Col {col} not a formula"
    wb.close()


def test_export_excel_presupuesto_row_count(breakdown_result, excel_path):
    """Presupuesto sheet must have 18 data rows + header + total."""
    export_excel(breakdown_result, excel_path)
    wb = load_workbook(str(excel_path))
    ws = wb["Presupuesto"]
    assert ws.max_row == 20  # 1 header + 18 data + 1 total
    wb.close()


def test_export_excel_metadata_ids(breakdown_result, excel_path):
    """Metadata sheet must have all 4 isolation IDs."""
    export_excel(breakdown_result, excel_path)
    wb = load_workbook(str(excel_path))
    ws = wb["Metadata"]
    metadata = {}
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[0]:
            metadata[row[0]] = row[1]
    assert "organization_id" in metadata
    assert "tenant_id" in metadata
    assert "project_id" in metadata
    assert "film_id" in metadata
    assert metadata["organization_id"] == "ORG-DEMO-001"
    assert metadata["tenant_id"] == "TENANT-DEMO-001"
    assert metadata["project_id"] == "PROJECT-DEMO-001"
    assert metadata["film_id"] == "FILM-DEMO-001"
    wb.close()


def test_export_excel_is_editable(breakdown_result, excel_path):
    """Excel file must be openable and modifiable with openpyxl."""
    export_excel(breakdown_result, excel_path)
    wb = load_workbook(str(excel_path))
    ws = wb["Resumen"]
    ws.cell(row=1, column=3, value="Modified")
    assert ws.cell(row=1, column=3).value == "Modified"
    wb.close()


def test_export_excel_no_pdf_generated(breakdown_result, tmp_path):
    """Excel export must not generate .pdf files."""
    excel_path = tmp_path / "breakdown.xlsx"
    export_excel(breakdown_result, excel_path)
    pdf_files = list(tmp_path.rglob("*.pdf"))
    assert pdf_files == []


def test_export_excel_no_html_generated(breakdown_result, tmp_path):
    """Excel export must not generate .html files."""
    excel_path = tmp_path / "breakdown.xlsx"
    export_excel(breakdown_result, excel_path)
    html_files = list(tmp_path.rglob("*.html"))
    assert html_files == []


def test_export_excel_no_csv_generated(breakdown_result, tmp_path):
    """Excel export must not generate .csv files."""
    excel_path = tmp_path / "breakdown.xlsx"
    export_excel(breakdown_result, excel_path)
    csv_files = list(tmp_path.rglob("*.csv"))
    assert csv_files == []


def test_export_excel_wrong_extension(breakdown_result, tmp_path):
    """Excel export must reject wrong extension."""
    wrong_path = tmp_path / "breakdown.txt"
    with pytest.raises(ValueError, match="Extensión no permitida"):
        export_excel(breakdown_result, wrong_path)


def test_export_excel_header_style(breakdown_result, excel_path):
    """Header row must be bold."""
    export_excel(breakdown_result, excel_path)
    wb = load_workbook(str(excel_path))
    ws = wb["Escenas"]
    for col in range(1, 10):
        cell = ws.cell(row=1, column=col)
        assert cell.font.bold, f"Header cell ({1},{col}) not bold"
    wb.close()
