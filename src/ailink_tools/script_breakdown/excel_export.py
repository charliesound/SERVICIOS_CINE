"""Excel export for Script-to-Production Breakdown demo."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .schemas import BreakdownResult

ALLOWED_EXTENSIONS = {".xlsx"}

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
TRAFFIC_FILLS = {
    "verde": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
    "amarillo": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
    "naranja": PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),
    "rojo": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
}
TRAFFIC_FONTS = {
    "verde": Font(bold=True, color="FFFFFF"),
    "amarillo": Font(bold=True, color="000000"),
    "naranja": Font(bold=True, color="FFFFFF"),
    "rojo": Font(bold=True, color="FFFFFF"),
}


def _validate_path(path: Path) -> None:
    """Validate that the path has an allowed extension."""
    if not path.name:
        raise ValueError("Ruta vacía no permitida")
    if path.suffix.lower() not in ALLOWED_EXTENSIONS:
        raise ValueError(
            f"Extensión no permitida: {path.suffix}. "
            f"Permitidas: {ALLOWED_EXTENSIONS}"
        )


def _style_header(ws, col_count: int) -> None:
    """Apply header style to first row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws) -> None:
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)


def _add_alt_rows(ws, start_row: int = 2) -> None:
    """Apply alternating row colors."""
    for row in range(start_row, ws.max_row + 1):
        if (row - start_row) % 2 == 1:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = ALT_FILL


def _write_sheet(ws, headers: list[str], rows: list[list]) -> None:
    """Write headers and rows to a worksheet."""
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    _style_header(ws, len(headers))
    _auto_width(ws)
    _add_alt_rows(ws)
    ws.freeze_panes = "A2"


def _build_resumen(result: BreakdownResult) -> list[list]:
    """Build Resumen sheet data."""
    v = result.viability
    return [
        [
            "Título", result.project["title"],
        ],
        [
            "Tipo", result.project["type"],
        ],
        [
            "Género", result.project["genre"],
        ],
        [
            "Duración (min)", result.project["duration_minutes"],
        ],
        [
            "Semanas de rodaje", result.project["shooting_weeks"],
        ],
        [
            "Moneda", result.project["currency"],
        ],
        [
            "Viabilidad global", f"{v['global_score']}/{10}",
        ],
        [
            "Semáforo", v["global_traffic_light"],
        ],
        [
            "Resumen viabilidad", v["summary"],
        ],
    ]


def _build_escenas(result: BreakdownResult) -> list[list]:
    """Build Escenas sheet rows."""
    rows = []
    for s in result.scenes:
        rows.append([
            s.scene_id,
            s.number,
            s.header,
            s.location,
            s.int_ext,
            s.day_night,
            ", ".join(s.characters),
            s.complexity,
            s.notes,
        ])
    return rows


def _build_personajes(result: BreakdownResult) -> list[list]:
    """Build Personajes sheet rows."""
    rows = []
    for c in result.characters:
        rows.append([
            c.character_id,
            c.name,
            c.role,
            ", ".join(str(x) for x in c.scenes),
            c.age,
            c.complexity,
            c.notes,
        ])
    return rows


def _build_localizaciones(result: BreakdownResult) -> list[list]:
    """Build Localizaciones sheet rows."""
    rows = []
    for loc in result.locations:
        rows.append([
            loc.location_id,
            loc.name,
            loc.type,
            loc.int_ext,
            ", ".join(str(x) for x in loc.scenes),
            loc.permits,
            loc.complexity,
        ])
    return rows


def _build_riesgos(result: BreakdownResult) -> list[list]:
    """Build Riesgos sheet rows."""
    rows = []
    for r in result.risks:
        rows.append([
            r.risk_id,
            r.description,
            r.impact,
            r.probability,
            r.mitigation,
        ])
    return rows


def _build_viabilidad(result: BreakdownResult) -> list[list]:
    """Build Viabilidad sheet rows."""
    rows = []
    for ind in result.viability["indicators"]:
        rows.append([
            ind["indicator"],
            ind["score"],
            ind["max_score"],
            ind["traffic_light"],
            ind["justification"],
            ind["recommendation"],
        ])
    return rows


def _build_presupuesto(result: BreakdownResult) -> list[list]:
    """Build Presupuesto sheet rows."""
    rows = []
    for b in result.preliminary_budget:
        rows.append([
            b.budget_id,
            b.category,
            b.low,
            b.mid,
            b.high,
            b.confidence,
            b.assumptions,
        ])
    return rows


def _build_recomendaciones(result: BreakdownResult) -> list[list]:
    """Build Recomendaciones sheet rows."""
    rows = []
    for i, rec in enumerate(result.recommendations, 1):
        rows.append([i, rec])
    return rows


def _build_revision_humana(result: BreakdownResult) -> list[list]:
    """Build Revisión humana sheet rows."""
    rows = []
    for note in result.human_review_notes:
        rows.append([note])
    return rows


def _build_metadata(result: BreakdownResult) -> list[list]:
    """Build Metadata sheet rows."""
    m = result.metadata
    return [
        ["organization_id", m["organization_id"]],
        ["tenant_id", m["tenant_id"]],
        ["project_id", m["project_id"]],
        ["film_id", m["film_id"]],
        ["productora", m["organization_name"]],
        ["parser_version", m["parser_version"]],
        ["parser_type", m["parser_type"]],
        ["is_demo", m["is_demo"]],
    ]


def _add_presupuesto_total(ws, row_start: int) -> None:
    """Add TOTAL row with SUM formulas to Presupuesto sheet."""
    total_row = row_start + 1
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=2, value="")
    ws.cell(row=total_row, column=2).font = Font(bold=True)

    # SUM formulas for Baja (col 3), Media (col 4), Alta (col 5)
    for col in [3, 4, 5]:
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}2:{col_letter}{row_start})"
        cell = ws.cell(row=total_row, column=col, value=formula)
        cell.font = Font(bold=True)
        cell.number_format = "#,##0"

    # Style total row
    for col in range(1, 8):
        cell = ws.cell(row=total_row, column=col)
        if not cell.font.bold:
            cell.font = Font(bold=True)


def export_excel(result: BreakdownResult, path: Path) -> Path:
    """Export breakdown result to Excel file.

    Args:
        result: The breakdown result to export.
        path: The output file path (must end in .xlsx).

    Returns:
        The path to the created file.

    Raises:
        ValueError: If the path is invalid or has wrong extension.
    """
    _validate_path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # 1. Resumen
    ws = wb.active
    ws.title = "Resumen"
    resumen_data = _build_resumen(result)
    _write_sheet(ws, ["Campo", "Valor"], resumen_data)

    # 2. Escenas
    ws_escenas = wb.create_sheet("Escenas")
    _write_sheet(
        ws_escenas,
        ["ID", "Número", "Header", "Localización", "INT/EXT", "Día/Noche",
         "Personajes", "Complejidad", "Notas"],
        _build_escenas(result),
    )

    # 3. Personajes
    ws_personajes = wb.create_sheet("Personajes")
    _write_sheet(
        ws_personajes,
        ["ID", "Nombre", "Rol", "Escenas", "Edad", "Complejidad", "Notas"],
        _build_personajes(result),
    )

    # 4. Localizaciones
    ws_localizaciones = wb.create_sheet("Localizaciones")
    _write_sheet(
        ws_localizaciones,
        ["ID", "Nombre", "Tipo", "INT/EXT", "Escenas", "Permisos",
         "Complejidad"],
        _build_localizaciones(result),
    )

    # 5. Riesgos
    ws_riesgos = wb.create_sheet("Riesgos")
    _write_sheet(
        ws_riesgos,
        ["ID", "Descripción", "Impacto", "Probabilidad", "Mitigación"],
        _build_riesgos(result),
    )

    # 6. Viabilidad
    ws_viabilidad = wb.create_sheet("Viabilidad")
    viabilidad_rows = _build_viabilidad(result)
    _write_sheet(
        ws_viabilidad,
        ["Indicador", "Puntuación", "Máximo", "Semáforo", "Justificación",
         "Recomendación"],
        viabilidad_rows,
    )
    # Apply traffic light colors
    for row_idx in range(2, ws_viabilidad.max_row + 1):
        sem_cell = ws_viabilidad.cell(row=row_idx, column=4)
        light = str(sem_cell.value).lower() if sem_cell.value else ""
        if light in TRAFFIC_FILLS:
            sem_cell.fill = TRAFFIC_FILLS[light]
            sem_cell.font = TRAFFIC_FONTS[light]

    # 7. Presupuesto
    ws_presupuesto = wb.create_sheet("Presupuesto")
    presupuesto_rows = _build_presupuesto(result)
    _write_sheet(
        ws_presupuesto,
        ["ID", "Categoría", "Baja", "Media", "Alta", "Confianza",
         "Supuestos"],
        presupuesto_rows,
    )
    # Format budget columns as numbers
    for row in range(2, ws_presupuesto.max_row + 1):
        for col in [3, 4, 5]:
            cell = ws_presupuesto.cell(row=row, column=col)
            if cell.value is not None:
                cell.number_format = "#,##0"
    # Add TOTAL row
    _add_presupuesto_total(ws_presupuesto, len(presupuesto_rows) + 1)

    # 8. Recomendaciones
    ws_recomendaciones = wb.create_sheet("Recomendaciones")
    _write_sheet(
        ws_recomendaciones,
        ["Número", "Recomendación"],
        _build_recomendaciones(result),
    )

    # 9. Revisión humana
    ws_revision = wb.create_sheet("Revisión humana")
    _write_sheet(
        ws_revision,
        ["Nota"],
        _build_revision_humana(result),
    )

    # 10. Metadata
    ws_metadata = wb.create_sheet("Metadata")
    _write_sheet(
        ws_metadata,
        ["Campo", "Valor"],
        _build_metadata(result),
    )

    wb.save(str(path))
    return path
